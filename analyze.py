from openpyxl import load_workbook
from openpyxl import Workbook

# 新建一个新表用于存放结果数据
def new_sheet():
    workbook = load_workbook(filename="analyzedata.xlsx")
    # 定义一个新表名
    new_sheetname = "Result"
    # 判断Excel中是否已经存在“Result”名称的这样一个表
    if new_sheetname not in workbook.sheetnames:
        # 如果不存在我们就新建一个
        workbook.create_sheet(new_sheetname)
        # print(workbook.sheetnames)
        workbook.save(filename="analyzedata.xlsx")
    # 如果已经存在，我们就把sheet里的数据删除
    else:
        sheet = workbook[new_sheetname]
        # 打印出表格有数据的范围，观察看看
        print(sheet.dimensions) # A1:E27
        for row in sheet.iter_rows():
            print(row)
            sheet.delete_rows(idx=1)
            workbook.save(filename="analyzedata.xlsx")


# 到源数据表中找到目标数据
def find_result():
    data_cell_list = []
    workbookT = load_workbook(filename="analyzedata.xlsx")
    # sheetT = workbook.active
    sheetT = workbookT["SourceData"]
    # 获取工作表大小
    sheet_size = sheetT.dimensions
    cells = sheetT[sheet_size]
    # print(cells)
    data_cps_h31 = []
    data_cps_h32 = []
    data_poi_tp1 = []

    # cell_row_tuple是每一行为1个元组
    for cell_row_tuple in cells:
        # cell是每一行元组中的每一个小格子
        for cell in cell_row_tuple:
            # print(cell.value)
            if cell.value == "CPS_H31":
                print(cell.row, cell.column)
                cell_Y = sheetT.cell(row=(cell.row + 2), column=(cell.column + 3))
                cell_Z = sheetT.cell(row=(cell.row + 3), column=(cell.column + 3))
                cell_Len = sheetT.cell(row=(cell.row + 4), column=(cell.column + 3))
                cell_WID = sheetT.cell(row=(cell.row + 5), column=(cell.column + 3))
                cell_len_tem = sheetT.cell(row=(cell.row + 4), column=cell.column + 1)
                cell_wid_tem = sheetT.cell(row=(cell.row + 5), column=cell.column + 1)
                data_cps_h31 = [
                    ["Y", cell_Y.value],
                    ["Z", cell_Z.value],
                    [cell_len_tem.value, cell_Len.value],
                    [cell_wid_tem.value, cell_WID.value],
                ]

            if cell.value == "CPS_H32":
                print(cell.row, cell.column)
                cell_Y = sheetT.cell(row=(cell.row + 2), column=(cell.column + 3))
                cell_Z = sheetT.cell(row=(cell.row + 3), column=(cell.column + 3))
                cell_Len = sheetT.cell(row=(cell.row + 4), column=(cell.column + 3))
                cell_WID = sheetT.cell(row=(cell.row + 5), column=(cell.column + 3))
                cell_len_tem = sheetT.cell(row=(cell.row + 4), column=cell.column + 1)
                cell_wid_tem = sheetT.cell(row=(cell.row + 5), column=cell.column + 1)
                data_cps_h32 = [
                    ["Y", cell_Y.value],
                    ["Z", cell_Z.value],
                    [cell_len_tem.value, cell_Len.value],
                    [cell_wid_tem.value, cell_WID.value],
                ]

            if cell.value == "POI_TP1":
                print(cell.row, cell.column)
                cell_X = sheetT.cell(row=(cell.row + 2), column=(cell.column + 2))
                data_poi_tp1 = [
                    ["X", cell_X.value]
                ]

    data_re = [data_cps_h31, data_cps_h32, data_poi_tp1]
    return data_re

# 往表中添加数据
def append_rows(data_result):
    workbook = load_workbook(filename="analyzedata.xlsx")
    sheet = workbook["Result"]
    for row in data_result:
        sheet.append(row)

    workbook.save(filename="analyzedata.xlsx")

# 往sheet中插入一列后写入相关数据
def insert_cols_data():
    workbook = load_workbook(filename="analyzedata.xlsx")
    sheet = workbook["Result"]
    sheet.insert_cols(idx=1)
    cellOne = sheet["A1"]
    cellOne.value = "H31"
    cellTwo = sheet["A5"]
    cellTwo.value = "H32"
    cellThree = sheet["A9"]
    cellThree.value = "H31"
    workbook.save(filename="analyzedata.xlsx")


# 新建表
new_sheet()

# 源数据中获取目标数据
data_cps_h31 = find_result()[0]
data_cps_h32 = find_result()[1]
data_poi_tp1 = find_result()[2]

# 存储数据
append_rows(data_cps_h31)
append_rows(data_cps_h32)
append_rows(data_poi_tp1)

# 插入数据
insert_cols_data()
